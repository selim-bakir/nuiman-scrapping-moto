"""Export Excel (.xlsx) du rapport quotidien — avec photos et mise en forme."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Report, ProductResult
from .report import _status, fmt_restock

SIZE_ORDER = ["3XS", "2XS", "XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL"]

_GREEN = PatternFill("solid", fgColor="C6EFCE")
_RED = PatternFill("solid", fgColor="FFC7CE")
_ORANGE = PatternFill("solid", fgColor="FFEB9C")
_GREY = PatternFill("solid", fgColor="F2F2F2")
_HEADER = PatternFill("solid", fgColor="1F3864")
_ZEBRA = PatternFill("solid", fgColor="EEF3FA")
_TITLE = PatternFill("solid", fgColor="2E5496")
_STATUS_FILL = {"ok": _GREEN, "partial": _ORANGE, "rupture": _RED, "error": _GREY}
_STATUS_LABEL = {"ok": "Complet", "partial": "Partiel", "rupture": "Rupture", "error": "Erreur"}
_STATUS_RANK = {"rupture": 0, "partial": 1, "ok": 2, "error": 3}
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _ordered_sizes(results) -> list[str]:
    seen = {s.size for r in results for s in r.sizes}
    known = [s for s in SIZE_ORDER if s in seen]
    return known + sorted(seen - set(known))


def _thumb(url: str | None) -> bytes | None:
    if not url:
        return None
    try:
        r = requests.get(url.replace("/rs580/", "/rs150/"), timeout=15)
        if r.status_code == 200 and r.content:
            return r.content
    except Exception:
        pass
    return None


def build_xlsx(report: Report, path: Path) -> Path:
    """Génère un classeur stylé : 1 ligne/casque, photo + 1 colonne/taille (ruptures en tête)."""
    results = sorted(
        report.results,
        key=lambda r: (_STATUS_RANK.get(_status(r), 9), r.gamme or "", r.color or ""),
    )
    sizes = _ordered_sizes(report.results)
    headers = ["Photo", "Gamme", "Coloris", "Prix", "Statut", *sizes, "Fiche"]
    ncol = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dispo casques"

    # Ligne 1 : titre.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(row=1, column=1, value="🏍️  Disponibilité casques Shoei — Motoblouz")
    t.font = Font(bold=True, size=15, color="FFFFFF")
    t.fill = _TITLE
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Ligne 2 : sous-titre (date + compteurs).
    ok = sum(1 for r in report.results if _status(r) == "ok")
    part = sum(1 for r in report.results if _status(r) == "partial")
    rupt = sum(1 for r in report.results if _status(r) == "rupture")
    sub = (
        f"{report.generated_at:%d/%m/%Y %H:%M}   ·   {len(report.results)} casques   ·   "
        f"🟢 {ok} complets   🟡 {part} partiels   🔴 {rupt} ruptures"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    s = ws.cell(row=2, column=1, value=sub)
    s.font = Font(size=10, color="333333")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Ligne 3 : en-têtes de colonnes.
    hdr = 3
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=hdr, column=col, value=label)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = _HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER

    # Données.
    img_refs = []  # garde les BytesIO vivants jusqu'au save
    size_base = 6  # 1re colonne de taille
    for idx, r in enumerate(results):
        row = hdr + 1 + idx
        ws.row_dimensions[row].height = 58
        zebra = _ZEBRA if idx % 2 else None
        size_map = {s.size: s for s in r.sizes}

        values = [
            "",
            r.gamme or "",
            r.color or "",
            r.price or "",
            _STATUS_LABEL.get(_status(r), ""),
        ]
        for sz in sizes:
            st = size_map.get(sz)
            if st is None:
                values.append("—")
            elif st.available:
                values.append("Dispo")
            elif st.restock:
                values.append(f"Réappro {fmt_restock(st.restock)}")
            else:
                values.append("Rupture")
        values.append("")  # Fiche (lien posé après)

        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _BORDER
            if zebra and col < size_base:
                c.fill = zebra

        # Statut coloré.
        ws.cell(row=row, column=5).fill = _STATUS_FILL.get(_status(r), _GREY)
        # Cellules tailles colorées.
        for i, sz in enumerate(sizes):
            cell = ws.cell(row=row, column=size_base + i)
            st = size_map.get(sz)
            if st is None:
                cell.fill = _GREY
            elif st.available:
                cell.fill = _GREEN
            elif st.restock:
                cell.fill = _ORANGE
            else:
                cell.fill = _RED
        # Lien fiche.
        link = ws.cell(row=row, column=ncol, value="Voir")
        link.hyperlink = r.url
        link.font = Font(color="2E5496", underline="single")
        if zebra:
            link.fill = zebra

        # Photo.
        data = _thumb(r.image)
        if data:
            try:
                bio = BytesIO(data)
                img_refs.append(bio)
                xi = XLImage(bio)
                xi.width, xi.height = 72, 72
                ws.add_image(xi, f"A{row}")
            except Exception:
                pass

    # Largeurs.
    widths = [12, 13, 22, 9, 10] + [12] * len(sizes) + [8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hdr + 1}"
    ws.sheet_view.showGridLines = False

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
